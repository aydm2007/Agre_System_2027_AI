from __future__ import annotations

import hashlib
import json
from dataclasses import dataclass
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.conf import settings
from django.core.files.base import ContentFile
from django.db import models, transaction
from django.db.models import Case, Count, Max, Q, Sum, Value, When
from django.db.models.functions import Coalesce
from django.utils import timezone
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework.test import APIRequestFactory, force_authenticate

from smart_agri.core.api.permissions import user_farm_ids
from smart_agri.core.api.reporting_support import _normalize_section_scope
from smart_agri.core.api.reporting import advanced_report
from smart_agri.core.models import (
    Activity,
    Asset,
    AsyncImportJob,
    AsyncReportRequest,
    Attachment,
    AttachmentLifecycleEvent,
    AuditLog,
    CropPlan,
    DailyLog,
    Farm,
    FuelConsumptionAlert,
    ItemInventory,
    LocationTreeStock,
    MaterialVarianceAlert,
    RemoteReviewEscalation,
    RemoteReviewLog,
    StockMovement,
    TreeStockEvent,
    VarianceAlert,
)
from smart_agri.core.models.farm import Location
from smart_agri.core.services.contract_operations_service import ContractOperationsService
from smart_agri.core.services.fixed_asset_workflow_service import FixedAssetWorkflowService
from smart_agri.core.services.fuel_reconciliation_service import FuelReconciliationService
from smart_agri.core.services.import_export_catalog import EXPORT_DEFINITIONS
from smart_agri.core.services.inventory_service import InventoryService
from smart_agri.core.services.mode_policy_service import resolve_farm_settings
from smart_agri.core.services.planning_import_service import PlanningImportService
from smart_agri.core.services.sensitive_audit import log_sensitive_mutation
from smart_agri.core.services.xlsx_workbook_service import XlsxWorkbookService
from smart_agri.finance.models import ApprovalRequest, ApprovalStageEvent
from smart_agri.finance.models_petty_cash import PettyCashRequest, PettyCashSettlement
from smart_agri.finance.models_supplier_settlement import SupplierSettlement
from smart_agri.inventory.models import Item, Unit


TEMPLATE_VERSION = "v1"
IMPORT_META_SHEET = "__meta"


@dataclass(frozen=True)
class TemplateDefinition:
    code: str
    title: str
    description: str
    mode_scope: str
    module: str
    worksheet_title: str
    headers: tuple[str, ...]
    requires_crop_plan: bool = False


@dataclass(frozen=True)
class ExportDefinition:
    code: str
    title: str
    description: str
    report_group: str
    mode_scope: str
    role_scope: str
    sensitivity_level: str
    default_column_profile: str = "default"
    allowed_formats: tuple[str, ...] = (
        AsyncReportRequest.FORMAT_XLSX,
        AsyncReportRequest.FORMAT_JSON,
    )
    ui_surface: str = "reports_hub"


class ImportExportPlatformService:
    EXPORT_TYPE_LABELS = {
        AsyncReportRequest.EXPORT_TYPE_ADVANCED_REPORT: "تقارير الأنشطة المتقدمة",
        AsyncReportRequest.EXPORT_TYPE_INVENTORY_BALANCE: "رصيد المخزون الحالي",
        AsyncReportRequest.EXPORT_TYPE_INVENTORY_MOVEMENTS: "حركة المخزون",
        AsyncReportRequest.EXPORT_TYPE_INVENTORY_LOW_STOCK: "المواد منخفضة الرصيد",
    }

    TEMPLATE_DEFINITIONS = {
        AsyncImportJob.TEMPLATE_INVENTORY_COUNT_SHEET: TemplateDefinition(
            code=AsyncImportJob.TEMPLATE_INVENTORY_COUNT_SHEET,
            title="ورقة جرد المخزون",
            description="قالب جرد فعلي للمخزون بالمواقع مع كمية فعلية محسوبة.",
            mode_scope="mode_aware_operational",
            module=AsyncImportJob.MODULE_INVENTORY,
            worksheet_title="الجرد",
            headers=("المادة", "المجموعة", "الموقع", "الكمية_الحالية", "الكمية_الفعلية", "ملاحظة"),
        ),
        AsyncImportJob.TEMPLATE_INVENTORY_OPERATIONAL_ADJUSTMENT: TemplateDefinition(
            code=AsyncImportJob.TEMPLATE_INVENTORY_OPERATIONAL_ADJUSTMENT,
            title="قالب تسوية تشغيلية للمخزون",
            description="إدخال زيادات/نواقص تشغيلية آمنة للمخزون.",
            mode_scope="mode_aware_operational",
            module=AsyncImportJob.MODULE_INVENTORY,
            worksheet_title="التسويات",
            headers=("المادة", "المجموعة", "الموقع", "الكمية_الفرق", "المرجع", "رقم_الدفعة", "تاريخ_الانتهاء", "ملاحظة"),
        ),
        AsyncImportJob.TEMPLATE_INVENTORY_OPENING_BALANCE: TemplateDefinition(
            code=AsyncImportJob.TEMPLATE_INVENTORY_OPENING_BALANCE,
            title="قالب أرصدة افتتاحية للمخزون",
            description="أرصدة افتتاحية مع تكلفة وحدة للمواقع المخزنية. STRICT فقط.",
            mode_scope="strict_only",
            module=AsyncImportJob.MODULE_INVENTORY,
            worksheet_title="الأرصدة_الافتتاحية",
            headers=("المادة", "المجموعة", "الموقع", "الكمية_الافتتاحية", "تكلفة_الوحدة", "رقم_الدفعة", "تاريخ_الانتهاء", "ملاحظة"),
        ),
        AsyncImportJob.TEMPLATE_INVENTORY_ITEM_MASTER: TemplateDefinition(
            code=AsyncImportJob.TEMPLATE_INVENTORY_ITEM_MASTER,
            title="قالب المواد الأساسية للمخزون",
            description="إنشاء أو تحديث مواد المخزون الأساسية. STRICT فقط.",
            mode_scope="strict_only",
            module=AsyncImportJob.MODULE_INVENTORY,
            worksheet_title="المواد_الأساسية",
            headers=("المادة", "المجموعة", "الوحدة_النصية", "رمز_الوحدة", "سعر_الوحدة", "حد_إعادة_الطلب", "تتبع_دفعات", "فترة_التحريم", "ملاحظة"),
        ),
    }

    TEMPLATE_DEFINITIONS.update(
        {
            AsyncImportJob.TEMPLATE_PLANNING_MASTER_SCHEDULE: TemplateDefinition(
                code=AsyncImportJob.TEMPLATE_PLANNING_MASTER_SCHEDULE,
                title="قالب الخطة الرئيسية",
                description="استيراد الخطة الرئيسية والبرنامج الموسمي من قالب Excel عربي RTL.",
                mode_scope="mode_aware_operational",
                module=AsyncImportJob.MODULE_PLANNING,
                worksheet_title="Planning_Master_Schedule",
                headers=("اسم_الخطة", "الموسم", "المحصول", "الموقع", "من", "إلى", "المساحة", "العملة", "ملاحظات"),
            ),
            AsyncImportJob.TEMPLATE_PLANNING_CROP_PLAN_STRUCTURE: TemplateDefinition(
                code=AsyncImportJob.TEMPLATE_PLANNING_CROP_PLAN_STRUCTURE,
                title="قالب الهيكل التشغيلي للخطة",
                description="استيراد الأنشطة المخططة والهيكل التشغيلي للخطة الزراعية من قالب Excel موحد.",
                mode_scope="mode_aware_operational",
                module=AsyncImportJob.MODULE_PLANNING,
                worksheet_title="Planning_CropPlan_Structure",
                headers=("رقم_النشاط", "اسم_النشاط", "التاريخ", "الساعات", "المرحلة", "ملاحظات"),
                requires_crop_plan=True,
            ),
            AsyncImportJob.TEMPLATE_PLANNING_CROP_PLAN_BUDGET: TemplateDefinition(
                code=AsyncImportJob.TEMPLATE_PLANNING_CROP_PLAN_BUDGET,
                title="قالب ميزانية الخطة",
                description="استيراد بنود الميزانية وافتراضات التكلفة المرتبطة بالخطة الزراعية. STRICT فقط.",
                mode_scope="strict_only",
                module=AsyncImportJob.MODULE_PLANNING,
                worksheet_title="Planning_CropPlan_Budget",
                headers=("رقم_النشاط", "اسم_النشاط", "فئة_الميزانية", "الكمية", "الوحدة", "السعر", "الإجمالي", "العملة"),
                requires_crop_plan=True,
            ),
        }
    )

    @classmethod
    def available_export_templates(cls, *, actor=None, farm_id=None, report_group=None, ui_surface=None):
        farm = None
        if actor is not None and farm_id not in (None, "", "all"):
            farm = cls._resolve_actor_farm(actor=actor, farm_id=farm_id)
        mode_context = cls._mode_context_for_farm(farm) if farm else ""
        results = []
        for export_type, definition in EXPORT_DEFINITIONS.items():
            if report_group and definition.report_group != report_group:
                continue
            if ui_surface and definition.ui_surface != ui_surface:
                continue
            if definition.mode_scope == "strict_only" and mode_context and mode_context != "STRICT":
                continue
            results.append(
                {
                    "export_type": export_type,
                    "title": definition.title,
                    "description": definition.description,
                    "formats": list(definition.allowed_formats),
                    "template_code": f"{export_type}_{TEMPLATE_VERSION}",
                    "template_version": TEMPLATE_VERSION,
                    "rtl": True,
                    "locale": "ar-YE",
                    "report_group": definition.report_group,
                    "mode_scope": definition.mode_scope,
                    "role_scope": definition.role_scope,
                    "sensitivity_level": definition.sensitivity_level,
                    "default_column_profile": definition.default_column_profile,
                    "ui_surface": definition.ui_surface,
                }
            )
        return results

    @classmethod
    def available_import_templates(cls, *, actor=None, farm_id=None, module=None):
        farm = None
        if actor is not None and farm_id not in (None, "", "all"):
            farm = cls._resolve_actor_farm(actor=actor, farm_id=farm_id)
        mode_context = cls._mode_context_for_farm(farm) if farm else ""
        templates = [
            {
                "code": definition.code,
                "title": definition.title,
                "description": definition.description,
                "mode_scope": definition.mode_scope,
                "module": definition.module,
                "template_version": TEMPLATE_VERSION,
                "format": AsyncReportRequest.FORMAT_XLSX,
                "farm_id": farm.id if farm else None,
                "mode_context": mode_context,
                "requires_crop_plan": definition.requires_crop_plan,
            }
            for definition in cls.TEMPLATE_DEFINITIONS.values()
        ]
        if module:
            templates = [entry for entry in templates if entry["module"] == module]
        if mode_context == "STRICT" or not mode_context:
            return templates
        return [entry for entry in templates if entry["mode_scope"] != "strict_only"]

    @classmethod
    def list_export_jobs(cls, *, actor, farm_id=None, limit=20):
        queryset = AsyncReportRequest.objects.filter(created_by=actor).order_by("-requested_at", "-id")
        if actor.is_superuser and farm_id not in (None, "", "all"):
            queryset = queryset.filter(params__farm_id=int(farm_id))
        elif farm_id not in (None, "", "all"):
            target_farm = cls._resolve_actor_farm(actor=actor, farm_id=farm_id)
            queryset = queryset.filter(params__farm_id=target_farm.id)
        return list(queryset[:limit])

    @classmethod
    def list_import_jobs(cls, *, actor, farm_id=None, limit=20, module=None):
        queryset = AsyncImportJob.objects.filter(created_by=actor).order_by("-requested_at", "-id")
        if actor.is_superuser and farm_id not in (None, "", "all"):
            queryset = queryset.filter(farm_id=int(farm_id))
        elif farm_id not in (None, "", "all"):
            target_farm = cls._resolve_actor_farm(actor=actor, farm_id=farm_id)
            queryset = queryset.filter(farm_id=target_farm.id)
        if module:
            queryset = queryset.filter(module=module)
        return list(queryset[:limit])

    @classmethod
    def create_export_job(cls, *, actor, payload: dict) -> AsyncReportRequest:
        farm = cls._resolve_actor_farm(actor=actor, farm_id=payload.get("farm_id"))
        export_type = payload.get("export_type") or AsyncReportRequest.EXPORT_TYPE_ADVANCED_REPORT
        output_format = payload.get("format") or AsyncReportRequest.FORMAT_XLSX
        export_definition = cls._get_export_definition(export_type)
        if output_format not in export_definition.allowed_formats:
            raise ValidationError("صيغة التصدير غير مدعومة لهذا التقرير.")
        mode_context = cls._mode_context_for_farm(farm)
        if export_definition.mode_scope == "strict_only" and mode_context != "STRICT":
            raise PermissionDenied("هذا التقرير متاح فقط في الوضع الصارم STRICT.")
        template_code = payload.get("template_code") or f"{export_type}_{TEMPLATE_VERSION}"
        params = dict(payload or {})
        params["farm_id"] = farm.id if farm else None
        section_scope = _normalize_section_scope(payload.get("section_scope"))
        job = AsyncReportRequest.objects.create(
            created_by=actor,
            report_type=AsyncReportRequest.REPORT_ADVANCED,
            export_type=export_type,
            output_format=output_format,
            template_code=template_code,
            template_version=payload.get("template_version") or TEMPLATE_VERSION,
            locale=payload.get("locale") or "ar-YE",
            rtl=bool(payload.get("rtl", True)),
            params=params,
            metadata={
                "mode_context": mode_context,
                "column_profile": payload.get("column_profile") or export_definition.default_column_profile,
                "report_group": export_definition.report_group,
                "sensitivity_level": export_definition.sensitivity_level,
                "role_scope": export_definition.role_scope,
                "ui_surface": export_definition.ui_surface,
                "section_scope": section_scope,
            },
        )
        log_sensitive_mutation(
            actor=actor,
            action="create_export_job",
            model_name="AsyncReportRequest",
            object_id=job.pk,
            reason=f"export:{export_type}",
            old_value=None,
            new_value={
                "export_type": export_type,
                "format": output_format,
                "template_code": template_code,
            },
            farm_id=farm.id if farm else None,
            context={"template_version": job.template_version},
        )
        return job

    @classmethod
    def generate_export_job(cls, job: AsyncReportRequest) -> None:
        job.mark_running()
        payload = cls._build_export_payload(job=job)
        filename, result_url = cls._persist_export_output(job=job, payload=payload)
        job.output_filename = filename
        job.mark_completed(result_url)
        job.save(update_fields=["output_filename"])

    @classmethod
    def build_template_workbook(cls, *, actor, farm_id: int, template_code: str, context: dict | None = None) -> bytes:
        context = context or {}
        farm = cls._resolve_actor_farm(actor=actor, farm_id=farm_id)
        template = cls._get_template(template_code)
        cls._enforce_template_mode(actor=actor, farm=farm, template=template, apply=False)
        crop_plan = None
        if template.requires_crop_plan:
            crop_plan = PlanningImportService._resolve_crop_plan_for_farm(
                farm=farm,
                crop_plan_id=context.get("crop_plan_id"),
                required=True,
            )
        workbook = XlsxWorkbookService.create_workbook()
        XlsxWorkbookService.add_cover_sheet(
            workbook,
            title=template.title,
            rows=[
                ("الوصف", template.description),
                ("المزرعة", farm.name),
                ("نطاق المود", template.mode_scope),
                ("الخطة المستهدفة", crop_plan.name if crop_plan else ""),
                ("الإصدار", TEMPLATE_VERSION),
            ],
        )
        rows = cls._build_template_rows(template=template, farm=farm, context=context)
        XlsxWorkbookService.add_table_sheet(
            workbook,
            title=template.worksheet_title,
            headers=list(template.headers),
            rows=rows,
        )
        meta_rows = [
            ["template_code", template.code],
            ["template_version", TEMPLATE_VERSION],
            ["module", template.module],
            ["farm_scope", str(farm.id)],
            ["mode_scope", template.mode_scope],
            ["crop_plan_id", str(crop_plan.id) if crop_plan else ""],
            ["generated_at", timezone.now().isoformat()],
            ["checksum", cls._template_checksum(template.code, farm.id, context=context)],
        ]
        XlsxWorkbookService.add_table_sheet(
            workbook,
            title=IMPORT_META_SHEET,
            headers=["المفتاح", "القيمة"],
            rows=meta_rows,
            hidden=True,
        )
        return XlsxWorkbookService.to_bytes(workbook)

    @classmethod
    def create_import_job(cls, *, actor, farm_id: int, template_code: str, upload, context: dict | None = None) -> AsyncImportJob:
        context = context or {}
        farm = cls._resolve_actor_farm(actor=actor, farm_id=farm_id)
        template = cls._get_template(template_code)
        cls._enforce_template_mode(actor=actor, farm=farm, template=template, apply=False)
        crop_plan = None
        if template.requires_crop_plan:
            crop_plan = PlanningImportService._resolve_crop_plan_for_farm(
                farm=farm,
                crop_plan_id=context.get("crop_plan_id"),
                required=True,
            )
        file_bytes = upload.read()
        upload.seek(0)
        job = AsyncImportJob.objects.create(
            created_by=actor,
            farm=farm,
            module=template.module,
            template_code=template.code,
            template_version=TEMPLATE_VERSION,
            mode_context=cls._mode_context_for_farm(farm),
            metadata={
                "file_hash": cls._hash_bytes(file_bytes),
                "mode_scope": template.mode_scope,
                "crop_plan_id": crop_plan.id if crop_plan else None,
            },
        )
        job.uploaded_file.save(upload.name, ContentFile(file_bytes), save=True)
        job.mark_uploaded(uploaded_filename=upload.name)
        log_sensitive_mutation(
            actor=actor,
            action="create_import_job",
            model_name="AsyncImportJob",
            object_id=job.pk,
            reason=f"import:{template.code}",
            old_value=None,
            new_value={
                "template_code": template.code,
                "template_version": TEMPLATE_VERSION,
                "uploaded_filename": upload.name,
            },
            farm_id=farm.id,
            context={"mode_scope": template.mode_scope, "crop_plan_id": crop_plan.id if crop_plan else None},
        )
        return job

    @classmethod
    def validate_import_job(cls, *, actor, job: AsyncImportJob):
        cls._assert_job_access(actor=actor, job=job)
        template = cls._get_template(job.template_code)
        cls._enforce_template_mode(actor=actor, farm=job.farm, template=template, apply=False)
        workbook = load_workbook(filename=job.uploaded_file.path, data_only=True)
        meta = cls._read_meta(workbook=workbook)
        cls._validate_template_meta(job=job, meta=meta)
        rows = cls._read_data_rows(workbook=workbook, template=template)
        preview_rows, summary = cls._validate_rows(job=job, template=template, rows=rows)
        if summary["errors"] > 0:
            job.error_workbook.save(
                f"errors-{job.id}.xlsx",
                ContentFile(cls._build_error_workbook(job=job, preview_rows=preview_rows)),
                save=True,
            )
        elif job.error_workbook:
            job.error_workbook.delete(save=False)
            job.error_workbook = None
            job.save(update_fields=["error_workbook"])
        job.mark_preview_ready(
            preview_rows=preview_rows,
            validation_summary=summary,
            row_count=summary["total_rows"],
            rejected_count=summary["errors"],
        )
        if summary["errors"] == 0:
            job.mark_approved_for_apply()
        return preview_rows, summary

    @classmethod
    @transaction.atomic
    def apply_import_job(cls, *, actor, job: AsyncImportJob):
        cls._assert_job_access(actor=actor, job=job)
        template = cls._get_template(job.template_code)
        cls._enforce_template_mode(actor=actor, farm=job.farm, template=template, apply=True)
        if job.status not in {AsyncImportJob.STATUS_APPROVED_FOR_APPLY, AsyncImportJob.STATUS_PREVIEW_READY}:
            raise ValidationError("يجب تنفيذ المعاينة والتحقق قبل التطبيق.")
        applied_count = 0
        rejected_count = 0
        for row in job.preview_rows or []:
            if row.get("severity") == "error":
                rejected_count += 1
                continue
            cls._apply_preview_row(actor=actor, job=job, template=template, row=row)
            applied_count += 1
        job.mark_applied(
            applied_count=applied_count,
            rejected_count=rejected_count,
            result_summary={
                "template_code": job.template_code,
                "applied_count": applied_count,
                "rejected_count": rejected_count,
            },
        )
        log_sensitive_mutation(
            actor=actor,
            action="apply_import_job",
            model_name="AsyncImportJob",
            object_id=job.pk,
            reason=f"bulk_import:{job.template_code}",
            old_value={"status": AsyncImportJob.STATUS_APPROVED_FOR_APPLY},
            new_value={"status": job.status, "applied_count": applied_count, "rejected_count": rejected_count},
            farm_id=job.farm_id,
            context={"template_code": job.template_code},
        )
        return job.result_summary

    @classmethod
    def _persist_export_output(cls, *, job: AsyncReportRequest, payload):
        reports_dir = Path(settings.MEDIA_ROOT) / "reports"  # agri-guardian: decimal-safe path join
        reports_dir.mkdir(parents=True, exist_ok=True)
        extension = job.output_format
        filename = f"{job.export_type}-{job.id}.{extension}"
        file_path = reports_dir / filename  # agri-guardian: decimal-safe path join
        if job.output_format == AsyncReportRequest.FORMAT_JSON:
            file_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
        elif job.output_format == AsyncReportRequest.FORMAT_XLSX:
            file_path.write_bytes(payload)
        else:
            raise ValidationError("Unsupported export format.")
        return filename, f"/media/reports/{filename}"

    @classmethod
    def _build_export_payload(cls, *, job: AsyncReportRequest):
        export_type = job.effective_export_type
        if export_type == AsyncReportRequest.EXPORT_TYPE_ADVANCED_REPORT:
            data = cls._build_advanced_report_payload(job=job)
            return data if job.output_format == AsyncReportRequest.FORMAT_JSON else cls._render_advanced_report_workbook(job=job, payload=data)
        builder_map = {
            AsyncReportRequest.EXPORT_TYPE_DAILY_EXECUTION_SUMMARY: cls._build_daily_execution_summary_payload,
            AsyncReportRequest.EXPORT_TYPE_DAILY_EXECUTION_DETAIL: cls._build_daily_execution_detail_payload,
            AsyncReportRequest.EXPORT_TYPE_PLAN_ACTUAL_VARIANCE: cls._build_plan_actual_variance_payload,
            AsyncReportRequest.EXPORT_TYPE_PERENNIAL_TREE_BALANCE: cls._build_perennial_tree_balance_payload,
            AsyncReportRequest.EXPORT_TYPE_OPERATIONAL_READINESS: cls._build_operational_readiness_payload,
            AsyncReportRequest.EXPORT_TYPE_INVENTORY_BALANCE: cls._build_inventory_export_payload,
            AsyncReportRequest.EXPORT_TYPE_INVENTORY_MOVEMENTS: cls._build_inventory_export_payload,
            AsyncReportRequest.EXPORT_TYPE_INVENTORY_LOW_STOCK: cls._build_inventory_export_payload,
            AsyncReportRequest.EXPORT_TYPE_INVENTORY_EXPIRY_BATCHES: cls._build_inventory_export_payload,
            AsyncReportRequest.EXPORT_TYPE_FUEL_POSTURE_REPORT: cls._build_fuel_posture_report_payload,
            AsyncReportRequest.EXPORT_TYPE_FIXED_ASSET_REGISTER: cls._build_fixed_asset_register_payload,
            AsyncReportRequest.EXPORT_TYPE_CONTRACT_OPERATIONS_POSTURE: cls._build_contract_operations_posture_payload,
            AsyncReportRequest.EXPORT_TYPE_SUPPLIER_SETTLEMENT_POSTURE: cls._build_supplier_settlement_posture_payload,
            AsyncReportRequest.EXPORT_TYPE_PETTY_CASH_POSTURE: cls._build_petty_cash_posture_payload,
            AsyncReportRequest.EXPORT_TYPE_RECEIPTS_DEPOSIT_POSTURE: cls._build_receipts_deposit_posture_payload,
            AsyncReportRequest.EXPORT_TYPE_GOVERNANCE_WORK_QUEUE: cls._build_governance_work_queue_payload,
        }
        builder = builder_map.get(export_type)
        if builder is not None:
            data = builder(job=job)
            return data if job.output_format == AsyncReportRequest.FORMAT_JSON else cls._render_structured_workbook(job=job, payload=data)
        raise ValidationError("Unsupported export type.")

    @classmethod
    def _build_advanced_report_payload(cls, *, job: AsyncReportRequest):
        factory = APIRequestFactory()
        request = factory.get("/api/v1/advanced-report/", data=job.params or {})
        force_authenticate(request, user=job.created_by)
        response = advanced_report(request)
        if response.status_code != 200:
            raise ValidationError(f"تعذر إنشاء التقرير المتقدم. status={response.status_code}")
        return response.data

    @classmethod
    def _build_inventory_export_payload(cls, *, job: AsyncReportRequest):
        params = job.params or {}
        farm = cls._resolve_actor_farm(actor=job.created_by, farm_id=params.get("farm_id"))
        item_id = params.get("item") or params.get("item_id")
        location_id = params.get("location") or params.get("location_id")
        export_type = job.effective_export_type
        if export_type == AsyncReportRequest.EXPORT_TYPE_INVENTORY_BALANCE:
            qs = ItemInventory.objects.filter(farm=farm, deleted_at__isnull=True).select_related("item", "location")
            if item_id:
                qs = qs.filter(item_id=item_id)
            if location_id:
                qs = qs.filter(location__isnull=True) if str(location_id).lower() == "null" else qs.filter(location_id=location_id)
            rows = [
                {
                    "المادة": record.item.name,
                    "المجموعة": record.item.group,
                    "الموقع": record.location.name if record.location else "بدون موقع",
                    "الكمية": str(record.qty),
                    "الوحدة": record.uom or record.item.uom,
                    "سعر_الوحدة": str(record.item.unit_price),
                    "حد_إعادة_الطلب": str(record.item.reorder_level or Decimal("0")),
                    "منخفض": "نعم" if record.qty < (record.item.reorder_level or Decimal("0")) else "لا",
                }
                for record in qs.order_by("item__group", "item__name")
            ]
            return {"title": "رصيد المخزون الحالي", "rows": rows, "filters": params}

        if export_type == AsyncReportRequest.EXPORT_TYPE_INVENTORY_LOW_STOCK:
            low_qs = ItemInventory.objects.filter(
                farm=farm,
                deleted_at__isnull=True,
                qty__lt=models.F("item__reorder_level"),
            ).select_related("item", "location")
            if item_id:
                low_qs = low_qs.filter(item_id=item_id)
            if location_id:
                low_qs = low_qs.filter(location__isnull=True) if str(location_id).lower() == "null" else low_qs.filter(location_id=location_id)
            rows = [
                {
                    "المادة": record.item.name,
                    "المجموعة": record.item.group,
                    "الموقع": record.location.name if record.location else "بدون موقع",
                    "الكمية": str(record.qty),
                    "حد_إعادة_الطلب": str(record.item.reorder_level or Decimal("0")),
                    "العجز": str((record.item.reorder_level or Decimal("0")) - record.qty),
                }
                for record in low_qs.order_by("item__name")
            ]
            return {"title": "المواد منخفضة الرصيد", "rows": rows, "filters": params}

        if export_type == AsyncReportRequest.EXPORT_TYPE_INVENTORY_EXPIRY_BATCHES:
            expiry_qs = (
                StockMovement.objects.filter(farm=farm)
                .exclude(batch_number__isnull=True)
                .exclude(batch_number="")
                .select_related("item", "location")
            )
            if item_id:
                expiry_qs = expiry_qs.filter(item_id=item_id)
            if location_id:
                expiry_qs = expiry_qs.filter(location__isnull=True) if str(location_id).lower() == "null" else expiry_qs.filter(location_id=location_id)
            rows = [
                {
                    "المادة": movement.item.name,
                    "المجموعة": movement.item.group,
                    "الموقع": movement.location.name if movement.location else "بدون موقع",
                    "رقم_الدفعة": movement.batch_number or "",
                    "تاريخ_الانتهاء": movement.expiry_date.isoformat() if movement.expiry_date else "",
                    "الكمية": str(movement.qty_delta),
                    "المرجع": movement.ref_id,
                    "نوع_المرجع": movement.ref_type,
                }
                for movement in expiry_qs.order_by("expiry_date", "item__name")[:5000]
            ]
            return {"title": "الدفعات والانتهاء المخزني", "rows": rows, "filters": params}

        movement_qs = StockMovement.objects.filter(farm=farm).select_related("item", "location")
        if item_id:
            movement_qs = movement_qs.filter(item_id=item_id)
        if location_id:
            movement_qs = movement_qs.filter(location__isnull=True) if str(location_id).lower() == "null" else movement_qs.filter(location_id=location_id)
        rows = [
            {
                "التاريخ": movement.created_at.date().isoformat() if movement.created_at else "",
                "المادة": movement.item.name,
                "المجموعة": movement.item.group,
                "الموقع": movement.location.name if movement.location else "بدون موقع",
                "الكمية": str(movement.qty_delta),
                "المرجع": movement.ref_id,
                "نوع_المرجع": movement.ref_type,
                "ملاحظة": movement.note,
                "رقم_الدفعة": movement.batch_number or "",
            }
            for movement in movement_qs.order_by("-created_at")[:5000]
        ]
        return {"title": "حركة المخزون", "rows": rows, "filters": params}

    @classmethod
    def _build_daily_execution_summary_payload(cls, *, job: AsyncReportRequest):
        params = job.params or {}
        farm = cls._resolve_actor_farm(actor=job.created_by, farm_id=params.get("farm_id"))
        queryset = (
            Activity.objects.filter(log__farm=farm, deleted_at__isnull=True)
            .select_related("log__farm", "crop", "task")
            .prefetch_related("activity_locations__location")
        )
        if params.get("start"):
            queryset = queryset.filter(log__log_date__gte=params["start"])
        if params.get("end"):
            queryset = queryset.filter(log__log_date__lte=params["end"])
        if params.get("crop_id") or params.get("crop"):
            queryset = queryset.filter(crop_id=params.get("crop_id") or params.get("crop"))
        rows = []
        for activity in queryset.order_by("-log__log_date", "id")[:5000]:
            location_names = [link.location.name for link in activity.activity_locations.all()]
            rows.append(
                {
                    "التاريخ": activity.log.log_date.isoformat() if activity.log_id else "",
                    "المزرعة": activity.log.farm.name if activity.log_id else "",
                    "المواقع": "، ".join(location_names) if location_names else "بدون موقع",
                    "المحصول": getattr(activity.crop, "name", ""),
                    "المهمة": getattr(activity.task, "name", ""),
                    "عدد_السجلات": 1,
                    "ساعات_العمل": str(activity.days_spent or Decimal("0")),
                    "ساعات_الآلة": str(activity.machine_hours or Decimal("0")),
                    "كمية_المياه": str(activity.water_volume or Decimal("0")),
                    "كمية_السماد": str(activity.fertilizer_quantity or Decimal("0")),
                }
            )
        return {"title": "ملخص التنفيذ اليومي", "rows": rows, "filters": params}

    @classmethod
    def _build_daily_execution_detail_payload(cls, *, job: AsyncReportRequest):
        payload = cls._build_advanced_report_payload(job=job)
        rows = []
        for entry in payload.get("details", []) or []:
            rows.append(
                {
                    "التاريخ": entry.get("log_date"),
                    "الموقع": ((entry.get("location") or {}).get("name") or ""),
                    "المحصول": ((entry.get("crop") or {}).get("name") or ""),
                    "المهمة": ((entry.get("task") or {}).get("name") or ""),
                    "المشرف": ((entry.get("supervisor") or {}).get("name") or ""),
                    "ساعات_العمل": entry.get("hours", 0),
                    "ساعات_الآلة": entry.get("machine_hours", 0),
                    "كمية_المياه": entry.get("water_volume", 0),
                    "كمية_السماد": entry.get("fertilizer_quantity", 0),
                    "كمية_الحصاد": entry.get("harvest_quantity", 0),
                }
            )
        return {"title": "سجل التنفيذ اليومي التفصيلي", "rows": rows, "filters": job.params or {}}

    @classmethod
    def _build_plan_actual_variance_payload(cls, *, job: AsyncReportRequest):
        params = job.params or {}
        farm = cls._resolve_actor_farm(actor=job.created_by, farm_id=params.get("farm_id"))
        plans = CropPlan.objects.filter(farm=farm, deleted_at__isnull=True).select_related("crop", "season")
        if params.get("crop_id") or params.get("crop"):
            plans = plans.filter(crop_id=params.get("crop_id") or params.get("crop"))
        rows = []
        for plan in plans.order_by("-created_at")[:500]:
            activities = Activity.objects.filter(crop_plan=plan, deleted_at__isnull=True)
            logs = DailyLog.objects.filter(activities__crop_plan=plan, deleted_at__isnull=True).distinct()
            open_variances = VarianceAlert.objects.filter(
                farm=farm,
                daily_log__in=logs,
                status__in=[VarianceAlert.ALERT_STATUS_UNINVESTIGATED, VarianceAlert.ALERT_STATUS_UNDER_REVIEW],
            )
            rows.append(
                {
                    "الخطة": getattr(plan, "name", "") or f"CropPlan #{plan.id}",
                    "المحصول": getattr(plan.crop, "name", ""),
                    "الموسم": getattr(plan.season, "name", ""),
                    "الحالة": getattr(plan, "status", ""),
                    "عدد_الأنشطة": activities.count(),
                    "عدد_اليوميات": logs.count(),
                    "ساعات_العمل_الفعلية": str(activities.aggregate(total=Coalesce(Sum("days_spent"), Value(Decimal("0"))))["total"]),
                    "ساعات_الآلة_الفعلية": str(activities.aggregate(total=Coalesce(Sum("machine_hours"), Value(Decimal("0"))))["total"]),
                    "انحرافات_مفتوحة": open_variances.count(),
                    "قيمة_الانحراف_المفتوح": str(open_variances.aggregate(total=Coalesce(Sum("variance_amount"), Value(Decimal("0"))))["total"]),
                }
            )
        return {"title": "الخطة مقابل الفعلي والانحراف", "rows": rows, "filters": params}

    @classmethod
    def _build_perennial_tree_balance_payload(cls, *, job: AsyncReportRequest):
        params = job.params or {}
        farm = cls._resolve_actor_farm(actor=job.created_by, farm_id=params.get("farm_id"))
        queryset = (
            LocationTreeStock.objects.filter(location__farm=farm, deleted_at__isnull=True)
            .select_related("location", "crop_variety", "crop_variety__crop", "productivity_status")
            .order_by("location__name", "crop_variety__crop__name", "crop_variety__name")
        )
        rows = []
        for stock in queryset[:5000]:
            events_qs = TreeStockEvent.objects.filter(stock=stock)
            rows.append(
                {
                    "الموقع": stock.location.name,
                    "المحصول": getattr(getattr(stock.crop_variety, "crop", None), "name", ""),
                    "الصنف": getattr(stock.crop_variety, "name", ""),
                    "الرصيد_الحالي": stock.current_tree_count,
                    "الحالة_الإنتاجية": getattr(stock.productivity_status, "name", ""),
                    "إجمالي_الأحداث": events_qs.count(),
                    "آخر_حدث": events_qs.order_by("-event_date").values_list("event_date", flat=True).first(),
                }
            )
        return {"title": "رصيد الأشجار والمعمّرات", "rows": rows, "filters": params}

    @classmethod
    def _build_operational_readiness_payload(cls, *, job: AsyncReportRequest):
        params = job.params or {}
        farm = cls._resolve_actor_farm(actor=job.created_by, farm_id=params.get("farm_id"))
        rows = [
            {"المؤشر": "اليوميات المتأخرة", "القيمة": DailyLog.objects.filter(farm=farm, status=DailyLog.STATUS_DRAFT, deleted_at__isnull=True).count(), "التصنيف": "readiness"},
            {"المؤشر": "انحرافات حرجة مفتوحة", "القيمة": VarianceAlert.objects.filter(farm=farm, status__in=[VarianceAlert.ALERT_STATUS_UNINVESTIGATED, VarianceAlert.ALERT_STATUS_UNDER_REVIEW]).count(), "التصنيف": "variance"},
            {"المؤشر": "تنبيهات وقود حرجة", "القيمة": FuelConsumptionAlert.objects.filter(log__farm=farm, status=FuelConsumptionAlert.STATUS_CRITICAL).count(), "التصنيف": "fuel"},
            {"المؤشر": "مرفقات تشغيلية بلا دورة حياة", "القيمة": Attachment.objects.filter(farm=farm, deleted_at__isnull=True).exclude(lifecycle_events__isnull=False).count(), "التصنيف": "attachments"},
            {"المؤشر": "انحرافات مواد مفتوحة", "القيمة": MaterialVarianceAlert.objects.filter(log__farm=farm, resolved_at__isnull=True).count(), "التصنيف": "materials"},
        ]
        return {"title": "جاهزية التشغيل", "rows": rows, "filters": params}

    @classmethod
    def _build_fuel_posture_report_payload(cls, *, job: AsyncReportRequest):
        params = job.params or {}
        payload = FuelReconciliationService.build_dashboard_payload(
            user=job.created_by,
            farm_id=params.get("farm_id"),
            tank=params.get("tank"),
        )
        rows = []
        for result in payload.get("results", []) or []:
            rows.append(
                {
                    "الأصل": result.get("asset_name") or result.get("asset", {}).get("name", ""),
                    "الحالة": result.get("status", ""),
                    "اللترات_الفعلية": result.get("actual_liters"),
                    "اللترات_المتوقعة": result.get("expected_liters"),
                    "الانحراف_لتر": result.get("variance_liters"),
                    "الانحراف_نسبة": result.get("variance_percentage"),
                    "جاهز_للاعتماد": result.get("approval_state", ""),
                }
            )
        return {"title": "وضعية الوقود والانحراف", "rows": rows, "filters": params}

    @classmethod
    def _build_fixed_asset_register_payload(cls, *, job: AsyncReportRequest):
        params = job.params or {}
        payload = FixedAssetWorkflowService.build_dashboard_payload(
            user=job.created_by,
            farm_id=params.get("farm_id"),
            category=params.get("category"),
        )
        rows = []
        for result in payload.get("results", []) or []:
            rows.append(
                {
                    "الكود": result.get("code", ""),
                    "الأصل": result.get("name", ""),
                    "الفئة": result.get("category", ""),
                    "الحالة": result.get("status", ""),
                    "حالة_الرأسملة": result.get("capitalization_state", ""),
                    "الصحة": result.get("health_status", ""),
                    "نسبة_الإهلاك": result.get("depreciation_percentage", ""),
                    "القيمة_الدفترية": result.get("book_value", ""),
                }
            )
        return {"title": "سجل الأصول الثابتة", "rows": rows, "filters": params}

    @classmethod
    def _build_contract_operations_posture_payload(cls, *, job: AsyncReportRequest):
        params = job.params or {}
        payload = ContractOperationsService.build_dashboard(farm_id=params.get("farm_id"))
        rows = []
        for result in payload.get("results", []) or []:
            rows.append(
                {
                    "المتعاقد": result.get("farmer_name", ""),
                    "نوع_العقد": result.get("contract_type", ""),
                    "المحصول": result.get("crop_name", ""),
                    "الموسم": result.get("season_name", ""),
                    "الحالة": result.get("status", ""),
                    "وضعية_التقييم": result.get("touring_state", ""),
                    "وضعية_التسوية": result.get("settlement_state", ""),
                    "الخطورة": result.get("variance_severity", ""),
                }
            )
        return {"title": "وضعية العقود الزراعية", "rows": rows, "filters": params}

    @classmethod
    def _build_supplier_settlement_posture_payload(cls, *, job: AsyncReportRequest):
        params = job.params or {}
        farm = cls._resolve_actor_farm(actor=job.created_by, farm_id=params.get("farm_id"))
        rows = []
        queryset = SupplierSettlement.objects.filter(farm=farm, deleted_at__isnull=True).select_related("purchase_order")
        for settlement in queryset.order_by("-created_at")[:1000]:
            rows.append(
                {
                    "المورد": settlement.vendor_name,
                    "مرجع_الفاتورة": settlement.invoice_reference,
                    "الحالة": settlement.status,
                    "طريقة_الدفع": settlement.payment_method,
                    "الواجب_سداده": str(settlement.payable_amount),
                    "المدفوع": str(settlement.paid_amount),
                    "المتبقي": str(settlement.remaining_balance),
                    "وضعية_المطابقة": settlement.reconciliation_state,
                    "الخطورة": settlement.variance_severity,
                }
            )
        return {"title": "وضعية تسويات الموردين", "rows": rows, "filters": params}

    @classmethod
    def _build_petty_cash_posture_payload(cls, *, job: AsyncReportRequest):
        params = job.params or {}
        farm = cls._resolve_actor_farm(actor=job.created_by, farm_id=params.get("farm_id"))
        rows = []
        queryset = PettyCashRequest.objects.filter(farm=farm, deleted_at__isnull=True).select_related("requester")
        for request_obj in queryset.order_by("-created_at")[:1000]:
            settlement = getattr(request_obj, "settlement", None)
            rows.append(
                {
                    "الطالب": getattr(request_obj.requester, "username", ""),
                    "الوصف": request_obj.description,
                    "الحالة": request_obj.status,
                    "المبلغ": str(request_obj.amount),
                    "حالة_التسوية": settlement.status if settlement else "OPEN",
                    "إجمالي_المصروف": str(settlement.total_expenses) if settlement else "0.0000",
                    "المسترد": str(settlement.refund_amount) if settlement else "0.0000",
                }
            )
        return {"title": "وضعية السلف والعهد النقدية", "rows": rows, "filters": params}

    @classmethod
    def _build_receipts_deposit_posture_payload(cls, *, job: AsyncReportRequest):
        params = job.params or {}
        farm = cls._resolve_actor_farm(actor=job.created_by, farm_id=params.get("farm_id"))
        receipt_audits = AuditLog.objects.filter(
            farm=farm,
            action__in=["RECEIPT_COLLECTED", "RECEIPT_DEPOSITED", "RECEIPT_RECONCILED", "RECEIPT_ANOMALY_FLAGGED"],
        ).order_by("-timestamp")[:2000]
        rows = [
            {
                "التاريخ": audit.timestamp.isoformat() if audit.timestamp else "",
                "الإجراء": audit.action,
                "المرجع": audit.object_id,
                "السبب": audit.reason,
                "المستخدم": getattr(audit.actor, "username", ""),
            }
            for audit in receipt_audits
        ]
        return {"title": "وضعية التحصيل والإيداع", "rows": rows, "filters": params}

    @classmethod
    def _build_governance_work_queue_payload(cls, *, job: AsyncReportRequest):
        params = job.params or {}
        farm = cls._resolve_actor_farm(actor=job.created_by, farm_id=params.get("farm_id"))
        if cls._mode_context_for_farm(farm) != "STRICT":
            raise PermissionDenied("تقرير الحوكمة متاح فقط في STRICT.")
        approval_rows = list(
            ApprovalRequest.objects.filter(farm=farm, deleted_at__isnull=True)
            .values("module", "required_role")
            .annotate(
                total_requests=Count("id"),
                pending=Count("id", filter=Q(status=ApprovalRequest.STATUS_PENDING)),
                approved=Count("id", filter=Q(status=ApprovalRequest.STATUS_APPROVED)),
                rejected=Count("id", filter=Q(status=ApprovalRequest.STATUS_REJECTED)),
            )
            .order_by("module", "required_role")
        )
        rows = [
            {
                "المسار": row["module"],
                "الدور": row["required_role"],
                "إجمالي_الطلبات": row["total_requests"],
                "معلقة": row["pending"],
                "معتمدة": row["approved"],
                "مرفوضة": row["rejected"],
            }
            for row in approval_rows
        ]
        rows.extend(
            [
                {
                    "المسار": "remote_review",
                    "الدور": "sector",
                    "إجمالي_الطلبات": RemoteReviewLog.objects.filter(farm=farm).count(),
                    "معلقة": RemoteReviewEscalation.objects.filter(farm=farm, resolved_at__isnull=True).count(),
                    "معتمدة": ApprovalStageEvent.objects.filter(request__farm=farm, action_type=ApprovalStageEvent.ACTION_FINAL_APPROVED).count(),
                    "مرفوضة": ApprovalStageEvent.objects.filter(request__farm=farm, action_type=ApprovalStageEvent.ACTION_REJECTED).count(),
                },
                {
                    "المسار": "attachments",
                    "الدور": "evidence",
                    "إجمالي_الطلبات": Attachment.objects.filter(farm=farm, deleted_at__isnull=True).count(),
                    "معلقة": AttachmentLifecycleEvent.objects.filter(attachment__farm=farm, action=AttachmentLifecycleEvent.ACTION_SCAN_QUARANTINED).count(),
                    "معتمدة": AttachmentLifecycleEvent.objects.filter(attachment__farm=farm, action=AttachmentLifecycleEvent.ACTION_AUTHORITATIVE_MARKED).count(),
                    "مرفوضة": AttachmentLifecycleEvent.objects.filter(attachment__farm=farm, action=AttachmentLifecycleEvent.ACTION_PURGED).count(),
                },
            ]
        )
        return {"title": "طوابير الحوكمة والموافقات", "rows": rows, "filters": params}

    @classmethod
    def _render_advanced_report_workbook(cls, *, job: AsyncReportRequest, payload: dict) -> bytes:
        workbook = XlsxWorkbookService.create_workbook()
        filters = payload.get("summary", {}).get("filters", {})
        summary = payload.get("summary", {}) or {}
        metrics = summary.get("metrics", {}) or {}
        details = payload.get("details", []) or []
        section_scope = payload.get("section_scope") or job.metadata.get("section_scope") or ["summary"]
        XlsxWorkbookService.add_cover_sheet(
            workbook,
            title="تقرير الأنشطة المتقدمة",
            rows=[
                ("المزرعة", str((filters.get("farms") or [""])[0] or "")),
                ("من", str((summary.get("period") or {}).get("start") or "")),
                ("إلى", str((summary.get("period") or {}).get("end") or "")),
                ("صيغة التصدير", job.output_format.upper()),
                ("القالب", job.template_code or ""),
                ("الأقسام", "، ".join(section_scope)),
            ],
        )
        XlsxWorkbookService.add_table_sheet(
            workbook,
            title="ملخص",
            headers=["المؤشر", "القيمة"],
            rows=[
                ["إجمالي ساعات العمل", metrics.get("total_hours", 0)],
                ["ساعات الآليات", metrics.get("machine_hours", 0)],
                ["إجمالي المواد", metrics.get("materials_total_qty", 0)],
                ["إجمالي الحصاد", metrics.get("harvest_total_qty", 0)],
            ],
            widths=[28, 18],
        )
        if details:
            XlsxWorkbookService.add_table_sheet(
                workbook,
                title="الأنشطة",
                headers=["التاريخ", "الموقع", "المحصول", "المهمة", "ساعات العمل", "ساعات الآلة", "قراءة البئر"],
                rows=[
                    [
                        row.get("log_date"),
                        ((row.get("location") or {}).get("name") or ""),
                        ((row.get("crop") or {}).get("name") or ""),
                        ((row.get("task") or {}).get("name") or ""),
                        row.get("hours", 0),
                        row.get("machine_hours", 0),
                        row.get("well_reading", 0),
                    ]
                    for row in details
                ],
                widths=[14, 18, 18, 22, 14, 14, 14],
            )
        if summary.get("locations"):
            XlsxWorkbookService.add_table_sheet(
                workbook,
                title="المواقع",
                headers=["الموقع", "ساعات العمل", "الحصاد", "عدد الأنشطة"],
                rows=[
                    [
                        row.get("name", ""),
                        row.get("total_hours", 0),
                        row.get("harvest_total_qty", 0),
                        row.get("activities", 0),
                    ]
                    for row in summary.get("locations", [])
                ],
                widths=[22, 14, 14, 14],
            )
        if payload.get("risk_zone"):
            XlsxWorkbookService.add_table_sheet(
                workbook,
                title="منطقة المخاطر",
                headers=["الخطة", "المهمة", "التاريخ", "التكلفة الفعلية", "المتوسط", "الانحراف"],
                rows=[
                    [
                        row.get("crop_plan_name", ""),
                        row.get("task_name", ""),
                        row.get("date", ""),
                        row.get("cost_total", 0),
                        row.get("mean", 0),
                        row.get("deviation", 0),
                    ]
                    for row in payload.get("risk_zone", [])
                ],
                widths=[22, 20, 14, 16, 16, 16],
            )
        perennial = summary.get("perennial_insights") or {}
        if perennial.get("summary"):
            XlsxWorkbookService.add_table_sheet(
                workbook,
                title="الأشجار",
                headers=["الموقع", "الصنف", "الرصيد الحالي"],
                rows=[
                    [
                        ((row.get("location") or {}).get("name") or row.get("location_name") or ""),
                        ((row.get("crop_variety") or {}).get("name") or row.get("variety_name") or ""),
                        row.get("current_tree_count", 0),
                    ]
                    for row in perennial.get("summary", [])
                ],
                widths=[22, 22, 14],
            )
        if perennial.get("events"):
            XlsxWorkbookService.add_table_sheet(
                workbook,
                title="أحداث الأشجار",
                headers=["التاريخ", "نوع الحدث", "التغير"],
                rows=[
                    [
                        row.get("event_timestamp", ""),
                        row.get("event_type", ""),
                        row.get("tree_count_delta", 0),
                    ]
                    for row in perennial.get("events", [])
                ],
                widths=[20, 18, 14],
            )
        XlsxWorkbookService.add_table_sheet(
            workbook,
            title=IMPORT_META_SHEET,
            headers=["المفتاح", "القيمة"],
            rows=[
                [key, json.dumps(value, ensure_ascii=False, default=str)]
                for key, value in ({**(filters or {}), "section_scope": section_scope}).items()
            ],
            widths=[24, 48],
            hidden=True,
        )
        return XlsxWorkbookService.to_bytes(workbook)

    @classmethod
    def _render_inventory_workbook(cls, *, job: AsyncReportRequest, payload: dict) -> bytes:
        workbook = XlsxWorkbookService.create_workbook()
        XlsxWorkbookService.add_cover_sheet(
            workbook,
            title=payload["title"],
            rows=[
                ("صيغة التصدير", job.output_format.upper()),
                ("المزرعة", str((job.params or {}).get("farm_id") or "")),
                ("القالب", job.template_code or ""),
                ("تم الإنشاء في", timezone.now().isoformat()),
            ],
        )
        rows = payload.get("rows", [])
        headers = list(rows[0].keys()) if rows else ["ملاحظة"]
        XlsxWorkbookService.add_table_sheet(
            workbook,
            title="البيانات",
            headers=headers,
            rows=[list(row.values()) for row in rows] if rows else [["لا توجد بيانات"]],
        )
        XlsxWorkbookService.add_table_sheet(
            workbook,
            title=IMPORT_META_SHEET,
            headers=["المفتاح", "القيمة"],
            rows=[[key, json.dumps(value, ensure_ascii=False, default=str)] for key, value in (payload.get("filters") or {}).items()],
            widths=[24, 48],
            hidden=True,
        )
        return XlsxWorkbookService.to_bytes(workbook)

    @classmethod
    def _render_structured_workbook(cls, *, job: AsyncReportRequest, payload: dict) -> bytes:
        workbook = XlsxWorkbookService.create_workbook()
        definition = cls._get_export_definition(job.effective_export_type)
        XlsxWorkbookService.add_cover_sheet(
            workbook,
            title=payload.get("title") or definition.title,
            rows=[
                ("المجموعة", definition.report_group),
                ("نطاق المود", definition.mode_scope),
                ("نطاق الدور", definition.role_scope),
                ("الحساسية", definition.sensitivity_level),
                ("المزرعة", str((job.params or {}).get("farm_id") or "")),
                ("تم الإنشاء في", timezone.now().isoformat()),
            ],
        )
        rows = payload.get("rows", [])
        headers = list(rows[0].keys()) if rows else ["ملاحظة"]
        XlsxWorkbookService.add_table_sheet(
            workbook,
            title="البيانات",
            headers=headers,
            rows=[list(row.values()) for row in rows] if rows else [["لا توجد بيانات"]],
        )
        XlsxWorkbookService.add_table_sheet(
            workbook,
            title=IMPORT_META_SHEET,
            headers=["المفتاح", "القيمة"],
            rows=[
                ["export_type", job.effective_export_type],
                ["template_code", job.template_code],
                ["template_version", job.template_version],
                ["report_group", definition.report_group],
                ["mode_scope", definition.mode_scope],
                ["role_scope", definition.role_scope],
                ["filters", json.dumps(payload.get("filters") or {}, ensure_ascii=False, default=str)],
            ],
            widths=[24, 64],
            hidden=True,
        )
        return XlsxWorkbookService.to_bytes(workbook)

    @classmethod
    def _apply_preview_row(cls, *, actor, job: AsyncImportJob, template: TemplateDefinition, row: dict):
        farm = job.farm
        if template.code == AsyncImportJob.TEMPLATE_INVENTORY_COUNT_SHEET:
            delta = Decimal(str(row["delta"]))
            if delta == 0:
                return
            InventoryService.record_movement(
                farm=farm,
                item=Item.objects.get(pk=row["item_id"]),
                qty_delta=delta,
                location=Location.objects.get(pk=row["location_id"]) if row.get("location_id") else None,
                ref_type="IMPORT_COUNT_SHEET",
                ref_id=f"{job.id}:{row['excel_row']}",
                note=row.get("note") or "جرد مخزون عبر Excel",
                actor_user=actor,
            )
            return
        if template.code == AsyncImportJob.TEMPLATE_INVENTORY_OPERATIONAL_ADJUSTMENT:
            InventoryService.record_movement(
                farm=farm,
                item=Item.objects.get(pk=row["item_id"]),
                qty_delta=Decimal(str(row["qty_delta"])),
                location=Location.objects.get(pk=row["location_id"]) if row.get("location_id") else None,
                ref_type="IMPORT_OPERATIONAL_ADJUSTMENT",
                ref_id=row.get("reference") or f"{job.id}:{row['excel_row']}",
                note=row.get("note") or "تسوية تشغيلية عبر Excel",
                batch_number=row.get("batch_number") or None,
                expiry_date=cls._parse_optional_date(row.get("expiry_date")),
                actor_user=actor,
            )
            return
        if template.code == AsyncImportJob.TEMPLATE_INVENTORY_OPENING_BALANCE:
            InventoryService.process_grn(
                farm=farm,
                item=Item.objects.get(pk=row["item_id"]),
                location=Location.objects.get(pk=row["location_id"]),
                qty=Decimal(str(row["opening_qty"])),
                unit_cost=Decimal(str(row["unit_cost"])),
                ref_id=f"OPENING-{job.id}-{row['excel_row']}",
                batch_number=row.get("batch_number") or None,
                expiry_date=cls._parse_optional_date(row.get("expiry_date")),
                actor_user=actor,
            )
            return
        if template.code == AsyncImportJob.TEMPLATE_INVENTORY_ITEM_MASTER:
            unit = None
            unit_code = (row.get("unit_code") or "").strip()
            if unit_code:
                unit = Unit.objects.filter(code=unit_code, deleted_at__isnull=True).first()
            item, _created = Item.objects.update_or_create(
                name=row["item_name"],
                group=row["item_group"],
                defaults={
                    "uom": row.get("uom") or "Unit",
                    "unit": unit,
                    "unit_price": Decimal(str(row.get("unit_price") or "0")),
                    "reorder_level": Decimal(str(row.get("reorder_level") or "0")),
                    "requires_batch_tracking": bool(row.get("requires_batch_tracking")),
                    "phi_days": int(row.get("phi_days") or 0),
                },
            )
            log_sensitive_mutation(
                actor=actor,
                action="upsert_item_master",
                model_name="Item",
                object_id=item.pk,
                reason="inventory_item_master_import",
                old_value=None,
                new_value={"name": item.name, "group": item.group},
                farm_id=farm.id,
                context={"job_id": job.id},
            )
            return
        if template.module == AsyncImportJob.MODULE_PLANNING:
            PlanningImportService.apply_row(
                actor=actor,
                template_code=template.code,
                farm=farm,
                row=row,
                job_id=job.id,
                job_metadata=job.metadata or {},
            )
            return
        raise ValidationError("Unsupported import template.")

    @classmethod
    def _validate_rows(cls, *, job: AsyncImportJob, template: TemplateDefinition, rows: list[dict]):
        preview_rows = []
        summary = {"total_rows": 0, "errors": 0, "warnings": 0, "valid_rows": 0}
        seen_keys = set()
        for row in rows:
            summary["total_rows"] += 1
            preview = {"excel_row": row["excel_row"], "severity": "ok", "messages": []}
            try:
                normalized = cls._normalize_import_row(job=job, template=template, row=row)
                duplicate_key = normalized.pop("__duplicate_key__", "")
                if duplicate_key and duplicate_key in seen_keys:
                    raise ValidationError("الصف مكرر داخل نفس ملف الاستيراد.")
                if duplicate_key:
                    seen_keys.add(duplicate_key)
                preview.update(normalized)
            except ValidationError as exc:
                preview["severity"] = "error"
                detail = getattr(exc, "message_dict", None) or getattr(exc, "messages", None) or [str(exc)]
                preview["messages"] = detail if isinstance(detail, list) else [str(detail)]
                summary["errors"] += 1
            else:
                if preview.get("warning"):
                    preview["severity"] = "warning"
                    preview["messages"].append(preview["warning"])
                    summary["warnings"] += 1
                summary["valid_rows"] += 1
            preview_rows.append(preview)
        return preview_rows, summary

    @classmethod
    def _normalize_import_row(cls, *, job: AsyncImportJob, template: TemplateDefinition, row: dict):
        farm = job.farm
        if template.code == AsyncImportJob.TEMPLATE_INVENTORY_COUNT_SHEET:
            item = cls._resolve_item(row["المادة"], row["المجموعة"])
            location = cls._resolve_location(farm=farm, location_name=row["الموقع"])
            current_qty = cls._resolve_inventory_qty(farm=farm, item=item, location=location)
            counted_qty = cls._to_decimal(row["الكمية_الفعلية"], "الكمية_الفعلية", allow_negative=False)
            delta = counted_qty - current_qty
            return {
                "item_id": item.id,
                "item_name": item.name,
                "item_group": item.group,
                "location_id": location.id if location else None,
                "location_name": location.name if location else "بدون موقع",
                "current_qty": str(current_qty),
                "counted_qty": str(counted_qty),
                "delta": str(delta),
                "note": row.get("ملاحظة") or "",
                "warning": "لا يوجد فرق لتطبيقه." if delta == 0 else "",
            }
        if template.code == AsyncImportJob.TEMPLATE_INVENTORY_OPERATIONAL_ADJUSTMENT:
            item = cls._resolve_item(row["المادة"], row["المجموعة"])
            location = cls._resolve_location(farm=farm, location_name=row["الموقع"])
            qty_delta = cls._to_decimal(row["الكمية_الفرق"], "الكمية_الفرق", allow_negative=True)
            if qty_delta == 0:
                raise ValidationError("الكمية_الفرق يجب ألا تكون صفراً.")
            if qty_delta < 0:
                current_qty = cls._resolve_inventory_qty(farm=farm, item=item, location=location)
                if current_qty + qty_delta < 0:
                    raise ValidationError("الكمية المطلوبة تخرق قيد عدم السماح برصيد سلبي.")
            return {
                "item_id": item.id,
                "item_name": item.name,
                "item_group": item.group,
                "location_id": location.id if location else None,
                "location_name": location.name if location else "بدون موقع",
                "qty_delta": str(qty_delta),
                "reference": str(row.get("المرجع") or ""),
                "batch_number": str(row.get("رقم_الدفعة") or ""),
                "expiry_date": row.get("تاريخ_الانتهاء") or "",
                "note": row.get("ملاحظة") or "",
            }
        if template.code == AsyncImportJob.TEMPLATE_INVENTORY_OPENING_BALANCE:
            item = cls._resolve_item(row["المادة"], row["المجموعة"])
            location = cls._resolve_location(farm=farm, location_name=row["الموقع"], require_value=True)
            opening_qty = cls._to_decimal(row["الكمية_الافتتاحية"], "الكمية_الافتتاحية", allow_negative=False)
            unit_cost = cls._to_decimal(row["تكلفة_الوحدة"], "تكلفة_الوحدة", allow_negative=False)
            current_qty = cls._resolve_inventory_qty(farm=farm, item=item, location=location)
            return {
                "item_id": item.id,
                "item_name": item.name,
                "item_group": item.group,
                "location_id": location.id,
                "location_name": location.name,
                "opening_qty": str(opening_qty),
                "unit_cost": str(unit_cost),
                "batch_number": str(row.get("رقم_الدفعة") or ""),
                "expiry_date": row.get("تاريخ_الانتهاء") or "",
                "note": row.get("ملاحظة") or "",
                "warning": "يوجد رصيد حالي لهذا الصنف في الموقع." if current_qty > 0 else "",
            }
        if template.code == AsyncImportJob.TEMPLATE_INVENTORY_ITEM_MASTER:
            item_name = str(row["المادة"] or "").strip()
            item_group = str(row["المجموعة"] or "").strip() or "General"
            if not item_name:
                raise ValidationError("اسم المادة مطلوب.")
            unit_price = cls._to_decimal(row["سعر_الوحدة"], "سعر_الوحدة", allow_negative=False, blank_as_zero=True)
            reorder_level = cls._to_decimal(row["حد_إعادة_الطلب"], "حد_إعادة_الطلب", allow_negative=False, blank_as_zero=True)
            return {
                "item_name": item_name,
                "item_group": item_group,
                "uom": str(row.get("الوحدة_النصية") or "Unit"),
                "unit_code": str(row.get("رمز_الوحدة") or ""),
                "unit_price": str(unit_price),
                "reorder_level": str(reorder_level),
                "requires_batch_tracking": cls._to_bool(row.get("تتبع_دفعات")),
                "phi_days": int(row.get("فترة_التحريم") or 0),
                "note": row.get("ملاحظة") or "",
            }
        if template.module == AsyncImportJob.MODULE_PLANNING:
            result = PlanningImportService.normalize_row(
                template_code=template.code,
                farm=job.farm,
                row=row,
                job_metadata=job.metadata or {},
            )
            payload = dict(result.payload)
            payload["__duplicate_key__"] = result.duplicate_key
            return payload
        raise ValidationError("Unsupported template.")

    @classmethod
    def _read_data_rows(cls, *, workbook: Workbook, template: TemplateDefinition):
        if template.worksheet_title not in workbook.sheetnames:
            raise ValidationError("ورقة البيانات غير موجودة في الملف.")
        worksheet = workbook[template.worksheet_title]
        header_row = [str(cell.value or "").strip() for cell in worksheet[1]]
        expected_headers = list(template.headers)
        if header_row[: len(expected_headers)] != expected_headers:
            raise ValidationError("رؤوس الأعمدة لا تطابق القالب المعتمد.")
        rows = []
        for excel_row in range(2, worksheet.max_row + 1):
            values = [worksheet.cell(row=excel_row, column=index).value for index in range(1, len(expected_headers) + 1)]
            if all(value in (None, "") for value in values):
                continue
            row = {header: value for header, value in zip(expected_headers, values)}
            row["excel_row"] = excel_row
            rows.append(row)
        return rows

    @classmethod
    def _read_meta(cls, *, workbook: Workbook):
        if IMPORT_META_SHEET not in workbook.sheetnames:
            raise ValidationError("ورقة metadata غير موجودة.")
        worksheet = workbook[IMPORT_META_SHEET]
        payload = {}
        for excel_row in range(2, worksheet.max_row + 1):
            key = worksheet.cell(row=excel_row, column=1).value
            value = worksheet.cell(row=excel_row, column=2).value
            if key:
                payload[str(key)] = value
        return payload

    @classmethod
    def _validate_template_meta(cls, *, job: AsyncImportJob, meta: dict):
        if str(meta.get("template_code") or "") != job.template_code:
            raise ValidationError("template_code غير مطابق.")
        if str(meta.get("template_version") or "") != job.template_version:
            raise ValidationError("template_version غير مطابق.")
        if str(meta.get("farm_scope") or "") != str(job.farm_id):
            raise ValidationError("الملف لا يطابق نطاق المزرعة المطلوب.")
        if str(meta.get("mode_scope") or "") != str(job.metadata.get("mode_scope") or ""):
            raise ValidationError("نطاق المود في القالب لا يطابق المهمة الحالية.")
        if str(meta.get("crop_plan_id") or "") != str(job.metadata.get("crop_plan_id") or ""):
            raise ValidationError("القالب لا يطابق الخطة الزراعية المستهدفة.")
        if str(meta.get("checksum") or "") != cls._template_checksum(job.template_code, job.farm_id, context=job.metadata or {}):
            raise ValidationError("ملف القالب غير صالح أو تم تعديله خارج القالب المعتمد.")

    @classmethod
    def _build_error_workbook(cls, *, job: AsyncImportJob, preview_rows: list[dict]) -> bytes:
        workbook = XlsxWorkbookService.create_workbook()
        XlsxWorkbookService.add_cover_sheet(
            workbook,
            title="ملف أخطاء الاستيراد",
            rows=[
                ("القالب", job.template_code),
                ("المزرعة", str(job.farm.name if job.farm else "")),
                ("وقت الإنشاء", timezone.now().isoformat()),
            ],
        )
        rows = [
            [
                preview.get("excel_row"),
                preview.get("severity"),
                " | ".join(str(message) for message in (preview.get("messages") or [])),
            ]
            for preview in preview_rows
            if preview.get("severity") in {"error", "warning"}
        ]
        XlsxWorkbookService.add_table_sheet(
            workbook,
            title="الأخطاء",
            headers=["صف_Excel", "الحالة", "الرسالة"],
            rows=rows or [["-", "ok", "لا توجد أخطاء"]],
            widths=[12, 12, 60],
        )
        return XlsxWorkbookService.to_bytes(workbook)

    @classmethod
    def _build_template_rows(cls, *, template: TemplateDefinition, farm: Farm, context: dict | None = None):
        context = context or {}
        if template.code == AsyncImportJob.TEMPLATE_INVENTORY_COUNT_SHEET:
            inventories = ItemInventory.objects.filter(farm=farm, deleted_at__isnull=True).select_related("item", "location")
            return [
                [
                    inventory.item.name,
                    inventory.item.group,
                    inventory.location.name if inventory.location else "",
                    str(inventory.qty),
                    "",
                    "",
                ]
                for inventory in inventories.order_by("item__group", "item__name")
            ] or [["", "", "", "", "", ""]]
        if template.code == AsyncImportJob.TEMPLATE_INVENTORY_OPERATIONAL_ADJUSTMENT:
            return [["", "", "", "", "", "", "", ""]]
        if template.code == AsyncImportJob.TEMPLATE_INVENTORY_OPENING_BALANCE:
            return [["", "", "", "", "", "", "", ""]]
        if template.code == AsyncImportJob.TEMPLATE_INVENTORY_ITEM_MASTER:
            return [["", "General", "Unit", "", "", "", "لا", "0", ""]]
        if template.module == AsyncImportJob.MODULE_PLANNING:
            return PlanningImportService.build_template_rows(
                template_code=template.code,
                farm=farm,
                crop_plan_id=context.get("crop_plan_id"),
            )
        return [[""] * len(template.headers)]

    @classmethod
    def _resolve_actor_farm(cls, *, actor, farm_id):
        if farm_id in (None, "", "all"):
            available_ids = user_farm_ids(actor)
            farm = Farm.objects.filter(id__in=available_ids, deleted_at__isnull=True).first() if not actor.is_superuser else Farm.objects.filter(deleted_at__isnull=True).first()
        else:
            farm = Farm.objects.filter(pk=farm_id, deleted_at__isnull=True).first()
        if farm is None:
            raise ValidationError("تعذر تحديد المزرعة ضمن نطاق المستخدم.")
        if not actor.is_superuser and farm.id not in set(user_farm_ids(actor)):
            raise PermissionDenied("ليس لك صلاحية على هذه المزرعة.")
        return farm

    @classmethod
    def _mode_context_for_farm(cls, farm: Farm):
        settings_obj, _, _ = resolve_farm_settings(farm=farm)
        return getattr(settings_obj, "mode", "")

    @classmethod
    def _enforce_template_mode(cls, *, actor, farm: Farm, template: TemplateDefinition, apply: bool):
        settings_obj, _, _ = resolve_farm_settings(farm=farm)
        mode = getattr(settings_obj, "mode", "")
        if template.mode_scope == "strict_only" and mode != "STRICT":
            raise PermissionDenied("هذا القالب متاح فقط في الوضع الصارم STRICT.")
        if apply and template.mode_scope == "mode_aware_operational":
            return

    @classmethod
    def _get_template(cls, template_code: str) -> TemplateDefinition:
        try:
            return cls.TEMPLATE_DEFINITIONS[template_code]
        except KeyError as exc:
            raise ValidationError("قالب الاستيراد غير معروف.") from exc

    @classmethod
    def _get_export_definition(cls, export_type: str):
        definition = EXPORT_DEFINITIONS.get(export_type)
        if definition is not None:
            return definition
        title = cls.EXPORT_TYPE_LABELS.get(export_type)
        if title:
            return ExportDefinition(
                code=export_type,
                title=title,
                description=title,
                report_group="legacy",
                mode_scope="simple_strict",
                role_scope="farm_or_sector",
                sensitivity_level="operational",
            )
        raise ValidationError("نوع التصدير غير معروف.")

    @classmethod
    def _assert_job_access(cls, *, actor, job):
        if actor.is_superuser:
            return
        if job.created_by_id != actor.id:
            raise PermissionDenied("ليس لك صلاحية على هذه المهمة.")
        if job.farm_id and job.farm_id not in set(user_farm_ids(actor)):
            raise PermissionDenied("نطاق المزرعة غير مسموح.")

    @staticmethod
    def _hash_bytes(content: bytes) -> str:
        return hashlib.sha256(content).hexdigest()

    @staticmethod
    def _template_checksum(template_code: str, farm_id: int, context: dict | None = None) -> str:
        context = context or {}
        crop_plan_id = context.get("crop_plan_id") or ""
        return hashlib.sha256(f"{template_code}:{TEMPLATE_VERSION}:{farm_id}:{crop_plan_id}".encode("utf-8")).hexdigest()[:24]

    @staticmethod
    def _resolve_item(item_name, item_group):
        item = Item.objects.filter(name=str(item_name).strip(), group=str(item_group or "General").strip()).first()
        if item is None:
            raise ValidationError(f"المادة غير موجودة: {item_name} / {item_group}")
        return item

    @staticmethod
    def _resolve_location(*, farm: Farm, location_name, require_value=False):
        value = str(location_name or "").strip()
        if not value:
            if require_value:
                raise ValidationError("اسم الموقع مطلوب لهذا القالب.")
            return None
        location = Location.objects.filter(farm=farm, name=value, deleted_at__isnull=True).first()
        if location is None:
            raise ValidationError(f"الموقع غير موجود ضمن المزرعة: {value}")
        return location

    @staticmethod
    def _resolve_inventory_qty(*, farm: Farm, item: Item, location):
        inventory = ItemInventory.objects.filter(farm=farm, item=item, location=location).first()
        return Decimal(str(inventory.qty if inventory else "0"))

    @staticmethod
    def _to_decimal(value, field_name, *, allow_negative, blank_as_zero=False):
        if value in (None, ""):
            if blank_as_zero:
                return Decimal("0")
            raise ValidationError(f"{field_name} مطلوب.")
        try:
            result = Decimal(str(value))
        except (InvalidOperation, TypeError) as exc:
            raise ValidationError(f"{field_name} قيمة رقمية غير صالحة.") from exc
        if not allow_negative and result < 0:
            raise ValidationError(f"{field_name} لا يقبل قيماً سالبة.")
        return result

    @staticmethod
    def _to_bool(value):
        candidate = str(value or "").strip().lower()
        return candidate in {"1", "true", "yes", "نعم", "y"}

    @staticmethod
    def _parse_optional_date(value):
        if not value:
            return None
        if isinstance(value, date):
            return value
        return date.fromisoformat(str(value))
