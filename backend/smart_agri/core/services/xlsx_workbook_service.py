from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class XlsxWorkbookService:
    HEADER_FILL = PatternFill("solid", fgColor="0F766E")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    TITLE_FONT = Font(size=14, bold=True)
    SUBTITLE_FONT = Font(size=11, bold=True, color="475569")
    THIN_BORDER = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    @classmethod
    def create_workbook(cls) -> Workbook:
        workbook = Workbook()
        workbook.properties.creator = "AgriAsset"
        return workbook

    @classmethod
    def configure_sheet(cls, worksheet, *, title: str) -> None:
        worksheet.title = title[:31]
        worksheet.sheet_view.rightToLeft = True
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = "A1:A1"

    @classmethod
    def add_cover_sheet(cls, workbook: Workbook, *, title: str, rows: list[tuple[str, str]]) -> None:
        worksheet = workbook.active
        cls.configure_sheet(worksheet, title="غلاف")
        worksheet["A1"] = title
        worksheet["A1"].font = cls.TITLE_FONT
        worksheet["A2"] = "ملف أعمال عربي RTL"
        worksheet["A2"].font = cls.SUBTITLE_FONT
        worksheet.column_dimensions["A"].width = 28
        worksheet.column_dimensions["B"].width = 60
        for index, (label, value) in enumerate(rows, start=4):
            worksheet[f"A{index}"] = label
            worksheet[f"B{index}"] = value
            for cell in (worksheet[f"A{index}"], worksheet[f"B{index}"]):
                cell.border = cls.THIN_BORDER
                cell.alignment = Alignment(horizontal="right", vertical="center")

    @classmethod
    def add_table_sheet(
        cls,
        workbook: Workbook,
        *,
        title: str,
        headers: list[str],
        rows: list[list[object]],
        widths: list[int] | None = None,
        hidden: bool = False,
    ):
        worksheet = workbook.create_sheet()
        cls.configure_sheet(worksheet, title=title)
        for idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=idx, value=header)
            cell.fill = cls.HEADER_FILL
            cell.font = cls.HEADER_FONT
            cell.border = cls.THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row_index, row_values in enumerate(rows, start=2):
            for column_index, value in enumerate(row_values, start=1):
                cell = worksheet.cell(row=row_index, column=column_index, value=value)
                cell.border = cls.THIN_BORDER
                cell.alignment = Alignment(horizontal="right", vertical="center")
        worksheet.auto_filter.ref = f"A1:{get_column_letter(max(len(headers), 1))}{max(len(rows) + 1, 1)}"
        if widths:
            for idx, width in enumerate(widths, start=1):
                worksheet.column_dimensions[get_column_letter(idx)].width = width
        else:
            for idx, header in enumerate(headers, start=1):
                max_len = max([len(str(header))] + [len(str(row[idx - 1])) for row in rows if len(row) >= idx], default=10)
                worksheet.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 4, 14), 40)
        if hidden:
            worksheet.sheet_state = "hidden"
        return worksheet

    @classmethod
    def to_bytes(cls, workbook: Workbook) -> bytes:
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()
