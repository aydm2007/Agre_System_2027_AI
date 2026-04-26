from decimal import Decimal
from io import BytesIO
from uuid import uuid4

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import override_settings
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APITestCase

from smart_agri.core.models import (
    AsyncImportJob,
    AsyncReportRequest,
    Crop,
    CropPlan,
    CropPlanBudgetLine,
    Farm,
    FarmSettings,
    ItemInventory,
    Location,
    PlannedActivity,
    Season,
    StockMovement,
    Task,
)
from smart_agri.core.services.import_export_platform_service import ImportExportPlatformService
from smart_agri.inventory.models import Item


@override_settings(MEDIA_ROOT="test_media/import_export")
class ImportExportPlatformTests(APITestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_superuser(
            username=f"import-export-{uuid4().hex[:8]}",
            email="import-export@example.com",
            password="pass1234",
        )
        self.client.force_authenticate(self.user)

        self.simple_farm = Farm.objects.create(
            name=f"Simple Farm {uuid4().hex[:6]}",
            slug=f"simple-farm-{uuid4().hex[:6]}",
            region="North",
        )
        FarmSettings.objects.create(farm=self.simple_farm, mode=FarmSettings.MODE_SIMPLE)

        self.strict_farm = Farm.objects.create(
            name=f"Strict Farm {uuid4().hex[:6]}",
            slug=f"strict-farm-{uuid4().hex[:6]}",
            region="North",
        )
        FarmSettings.objects.create(farm=self.strict_farm, mode=FarmSettings.MODE_STRICT)

        self.item = Item.objects.create(
            name=f"سماد إثباتي {uuid4().hex[:5]}",
            group="Inputs",
            uom="kg",
            unit_price=Decimal("15.500"),
            reorder_level=Decimal("2.000"),
        )
        self.season = Season.objects.create(
            name=f"2026-{uuid4().hex[:4]}",
            start_date="2026-01-01",
            end_date="2026-12-31",
            is_active=True,
        )
        self.crop = Crop.objects.create(
            name=f"قمح إثباتي {uuid4().hex[:4]}",
            mode="Open",
        )
        self.simple_location = Location.objects.create(
            farm=self.simple_farm,
            name=f"حقل بسيط {uuid4().hex[:4]}",
        )
        self.strict_location = Location.objects.create(
            farm=self.strict_farm,
            name=f"حقل صارم {uuid4().hex[:4]}",
        )
        self.task = Task.objects.create(
            crop=self.crop,
            name=f"حراثة إثباتية {uuid4().hex[:4]}",
            stage="General",
        )
        self.simple_plan = CropPlan.objects.create(
            farm=self.simple_farm,
            crop=self.crop,
            season=self.season,
            name=f"خطة بسيطة {uuid4().hex[:4]}",
            start_date="2026-01-10",
            end_date="2026-03-10",
            currency="YER",
        )
        self.strict_plan = CropPlan.objects.create(
            farm=self.strict_farm,
            crop=self.crop,
            season=self.season,
            name=f"خطة صارمة {uuid4().hex[:4]}",
            start_date="2026-02-01",
            end_date="2026-04-30",
            currency="YER",
        )

    def _response_bytes(self, response):
        if hasattr(response, "streaming_content"):
            return b"".join(response.streaming_content)
        return response.content

    def _build_upload(self, *, farm, template_code, row_values, crop_plan_id=None):
        workbook_bytes = ImportExportPlatformService.build_template_workbook(
            actor=self.user,
            farm_id=farm.id,
            template_code=template_code,
            context={"crop_plan_id": crop_plan_id} if crop_plan_id else None,
        )
        workbook = load_workbook(BytesIO(workbook_bytes))
        worksheet = workbook[
            ImportExportPlatformService.TEMPLATE_DEFINITIONS[template_code].worksheet_title
        ]
        for column_index, value in enumerate(row_values, start=1):
            worksheet.cell(row=2, column=column_index, value=value)
        handle = BytesIO()
        workbook.save(handle)
        return SimpleUploadedFile(
            name=f"{template_code}.xlsx",
            content=handle.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_downloaded_template_is_xlsx_with_rtl_and_metadata(self):
        response = self.client.get(
            f"/api/v1/import-templates/{AsyncImportJob.TEMPLATE_INVENTORY_COUNT_SHEET}/download/",
            {"farm_id": self.simple_farm.id},
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        workbook = load_workbook(BytesIO(self._response_bytes(response)))
        self.assertIn("__meta", workbook.sheetnames)
        self.assertTrue(workbook["الجرد"].sheet_view.rightToLeft)
        meta_sheet = workbook["__meta"]
        meta = {
            str(meta_sheet.cell(row=row, column=1).value): meta_sheet.cell(row=row, column=2).value
            for row in range(2, meta_sheet.max_row + 1)
        }
        self.assertEqual(meta["template_code"], AsyncImportJob.TEMPLATE_INVENTORY_COUNT_SHEET)
        self.assertEqual(str(meta["farm_scope"]), str(self.simple_farm.id))
        self.assertEqual(meta["mode_scope"], "mode_aware_operational")

    def test_strict_only_template_is_blocked_in_simple_mode(self):
        response = self.client.get(
            f"/api/v1/import-templates/{AsyncImportJob.TEMPLATE_INVENTORY_OPENING_BALANCE}/download/",
            {"farm_id": self.simple_farm.id},
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_inventory_operational_adjustment_import_validates_and_applies(self):
        upload = self._build_upload(
            farm=self.simple_farm,
            template_code=AsyncImportJob.TEMPLATE_INVENTORY_OPERATIONAL_ADJUSTMENT,
            row_values=[
                self.item.name,
                self.item.group,
                "",
                "4.5",
                "REF-001",
                "",
                "",
                "تسوية إثباتية",
            ],
        )

        upload_response = self.client.post(
            "/api/v1/import-jobs/upload/",
            {
                "template_code": AsyncImportJob.TEMPLATE_INVENTORY_OPERATIONAL_ADJUSTMENT,
                "farm_id": self.simple_farm.id,
                "file": upload,
            },
            format="multipart",
        )
        self.assertEqual(upload_response.status_code, status.HTTP_201_CREATED)
        job_id = upload_response.data["id"]

        validate_response = self.client.post(f"/api/v1/import-jobs/{job_id}/validate/", {}, format="json")
        self.assertEqual(validate_response.status_code, status.HTTP_200_OK)
        self.assertEqual(validate_response.data["status"], AsyncImportJob.STATUS_APPROVED_FOR_APPLY)
        self.assertEqual(validate_response.data["validation_summary"]["errors"], 0)

        apply_response = self.client.post(f"/api/v1/import-jobs/{job_id}/apply/", {}, format="json")
        self.assertEqual(apply_response.status_code, status.HTTP_200_OK)
        self.assertEqual(apply_response.data["status"], AsyncImportJob.STATUS_APPLIED)

        inventory = ItemInventory.objects.get(farm=self.simple_farm, item=self.item, location__isnull=True)
        self.assertEqual(inventory.qty, Decimal("4.5000"))
        movement = StockMovement.objects.get(
            farm=self.simple_farm,
            item=self.item,
            ref_type="IMPORT_OPERATIONAL_ADJUSTMENT",
        )
        self.assertEqual(movement.qty_delta, Decimal("4.5000"))

    def test_inventory_balance_export_supports_json_and_xlsx(self):
        ItemInventory.objects.create(
            farm=self.strict_farm,
            item=self.item,
            qty=Decimal("9.000"),
            uom="kg",
        )

        json_job = ImportExportPlatformService.create_export_job(
            actor=self.user,
            payload={
                "farm_id": self.strict_farm.id,
                "export_type": AsyncReportRequest.EXPORT_TYPE_INVENTORY_BALANCE,
                "format": AsyncReportRequest.FORMAT_JSON,
            },
        )
        ImportExportPlatformService.generate_export_job(json_job)
        json_job.refresh_from_db()
        self.assertEqual(json_job.status, AsyncReportRequest.STATUS_COMPLETED)
        self.assertTrue(json_job.output_filename.endswith(".json"))

        xlsx_job = ImportExportPlatformService.create_export_job(
            actor=self.user,
            payload={
                "farm_id": self.strict_farm.id,
                "export_type": AsyncReportRequest.EXPORT_TYPE_INVENTORY_BALANCE,
                "format": AsyncReportRequest.FORMAT_XLSX,
            },
        )
        payload = ImportExportPlatformService._build_export_payload(job=xlsx_job)
        workbook = load_workbook(BytesIO(payload))
        self.assertIn("البيانات", workbook.sheetnames)
        self.assertTrue(workbook["البيانات"].sheet_view.rightToLeft)

    def test_export_templates_expose_registry_metadata_and_filter_strict_only_for_simple(self):
        response = self.client.get(
            "/api/v1/export-templates/",
            {
                "farm_id": self.simple_farm.id,
                "ui_surface": "reports_hub",
            },
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        results = response.data["results"]
        execution_report = next(
            entry
            for entry in results
            if entry["export_type"] == AsyncReportRequest.EXPORT_TYPE_DAILY_EXECUTION_SUMMARY
        )
        self.assertEqual(execution_report["report_group"], "execution")
        self.assertEqual(execution_report["mode_scope"], "simple_strict")
        self.assertIn(AsyncReportRequest.FORMAT_XLSX, execution_report["formats"])
        self.assertNotIn(
            AsyncReportRequest.EXPORT_TYPE_GOVERNANCE_WORK_QUEUE,
            {entry["export_type"] for entry in results},
        )

    def test_strict_only_export_is_blocked_in_simple_mode(self):
        response = self.client.post(
            "/api/v1/export-jobs/",
            {
                "farm_id": self.simple_farm.id,
                "export_type": AsyncReportRequest.EXPORT_TYPE_GOVERNANCE_WORK_QUEUE,
                "format": AsyncReportRequest.FORMAT_XLSX,
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_operational_readiness_export_generates_structured_xlsx(self):
        job = ImportExportPlatformService.create_export_job(
            actor=self.user,
            payload={
                "farm_id": self.simple_farm.id,
                "export_type": AsyncReportRequest.EXPORT_TYPE_OPERATIONAL_READINESS,
                "format": AsyncReportRequest.FORMAT_XLSX,
            },
        )

        payload = ImportExportPlatformService._build_export_payload(job=job)
        workbook = load_workbook(BytesIO(payload))
        self.assertIn("غلاف", workbook.sheetnames)
        self.assertIn("البيانات", workbook.sheetnames)
        self.assertTrue(workbook["البيانات"].sheet_view.rightToLeft)
