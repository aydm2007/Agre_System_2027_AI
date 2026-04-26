from decimal import Decimal
from io import BytesIO
from uuid import uuid4

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import override_settings
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APITestCase

from smart_agri.core.models import (
    AsyncImportJob,
    Crop,
    CropPlan,
    CropPlanBudgetLine,
    Farm,
    FarmSettings,
    Location,
    PlannedActivity,
    Season,
    Task,
)
from smart_agri.core.services.import_export_platform_service import ImportExportPlatformService


@override_settings(MEDIA_ROOT="test_media/import_export")
class PlanningImportPlatformTests(APITestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_superuser(
            username=f"planning-import-{uuid4().hex[:8]}",
            email="planning-import@example.com",
            password="pass1234",
        )
        self.client.force_authenticate(self.user)

        self.simple_farm = Farm.objects.create(
            name=f"Simple Planning Farm {uuid4().hex[:6]}",
            slug=f"simple-planning-farm-{uuid4().hex[:6]}",
            region="North",
        )
        FarmSettings.objects.create(farm=self.simple_farm, mode=FarmSettings.MODE_SIMPLE)

        self.strict_farm = Farm.objects.create(
            name=f"Strict Planning Farm {uuid4().hex[:6]}",
            slug=f"strict-planning-farm-{uuid4().hex[:6]}",
            region="North",
        )
        FarmSettings.objects.create(farm=self.strict_farm, mode=FarmSettings.MODE_STRICT)

        self.season = Season.objects.create(
            name=f"2026-{uuid4().hex[:4]}",
            start_date="2026-01-01",
            end_date="2026-12-31",
            is_active=True,
        )
        self.crop = Crop.objects.create(
            name=f"قمح تخطيطي {uuid4().hex[:4]}",
            mode="Open",
        )
        self.simple_location = Location.objects.create(
            farm=self.simple_farm,
            name=f"حقل بسيط {uuid4().hex[:4]}",
        )
        self.strict_location = Location.objects.create(
            farm=self.strict_farm,
            name=f"حقل صارم {uuid4().hex[:4]}",
        )
        self.task = Task.objects.create(
            crop=self.crop,
            name=f"حراثة تخطيطية {uuid4().hex[:4]}",
            stage="General",
        )
        self.simple_plan = CropPlan.objects.create(
            farm=self.simple_farm,
            crop=self.crop,
            season=self.season,
            name=f"خطة بسيطة {uuid4().hex[:4]}",
            start_date="2026-01-10",
            end_date="2026-03-10",
            currency="YER",
        )
        self.strict_plan = CropPlan.objects.create(
            farm=self.strict_farm,
            crop=self.crop,
            season=self.season,
            name=f"خطة صارمة {uuid4().hex[:4]}",
            start_date="2026-02-01",
            end_date="2026-04-30",
            currency="YER",
        )

    def _response_bytes(self, response):
        if hasattr(response, "streaming_content"):
            return b"".join(response.streaming_content)
        return response.content

    def _build_upload(self, *, farm, template_code, row_values, crop_plan_id=None):
        workbook_bytes = ImportExportPlatformService.build_template_workbook(
            actor=self.user,
            farm_id=farm.id,
            template_code=template_code,
            context={"crop_plan_id": crop_plan_id} if crop_plan_id else None,
        )
        workbook = load_workbook(BytesIO(workbook_bytes))
        worksheet = workbook[
            ImportExportPlatformService.TEMPLATE_DEFINITIONS[template_code].worksheet_title
        ]
        for column_index, value in enumerate(row_values, start=1):
            worksheet.cell(row=2, column=column_index, value=value)
        handle = BytesIO()
        workbook.save(handle)
        return SimpleUploadedFile(
            name=f"{template_code}.xlsx",
            content=handle.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_planning_templates_are_discoverable_by_module(self):
        response = self.client.get(
            "/api/v1/import-templates/",
            {"farm_id": self.simple_farm.id, "module": "planning"},
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        codes = {entry["code"] for entry in response.data["results"]}
        self.assertIn(AsyncImportJob.TEMPLATE_PLANNING_MASTER_SCHEDULE, codes)
        self.assertIn(AsyncImportJob.TEMPLATE_PLANNING_CROP_PLAN_STRUCTURE, codes)
        self.assertNotIn(AsyncImportJob.TEMPLATE_PLANNING_CROP_PLAN_BUDGET, codes)

    def test_planning_structure_template_contains_crop_plan_metadata(self):
        response = self.client.get(
            f"/api/v1/import-templates/{AsyncImportJob.TEMPLATE_PLANNING_CROP_PLAN_STRUCTURE}/download/",
            {"farm_id": self.strict_farm.id, "crop_plan_id": self.strict_plan.id},
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        workbook = load_workbook(BytesIO(self._response_bytes(response)))
        meta_sheet = workbook["__meta"]
        meta = {
            str(meta_sheet.cell(row=row, column=1).value): meta_sheet.cell(row=row, column=2).value
            for row in range(2, meta_sheet.max_row + 1)
        }
        self.assertEqual(meta["template_code"], AsyncImportJob.TEMPLATE_PLANNING_CROP_PLAN_STRUCTURE)
        self.assertEqual(str(meta["crop_plan_id"]), str(self.strict_plan.id))

    def test_planning_budget_template_is_blocked_in_simple_mode(self):
        response = self.client.get(
            f"/api/v1/import-templates/{AsyncImportJob.TEMPLATE_PLANNING_CROP_PLAN_BUDGET}/download/",
            {"farm_id": self.simple_farm.id, "crop_plan_id": self.simple_plan.id},
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_planning_master_schedule_import_validates_and_applies(self):
        upload = self._build_upload(
            farm=self.strict_farm,
            template_code=AsyncImportJob.TEMPLATE_PLANNING_MASTER_SCHEDULE,
            row_values=[
                "خطة تشغيل ربيعية",
                self.season.name,
                self.crop.name,
                self.strict_location.name,
                "2026-03-01",
                "2026-05-01",
                "12.5",
                "YER",
                "استيراد إثباتي",
            ],
        )

        upload_response = self.client.post(
            "/api/v1/import-jobs/upload/",
            {
                "template_code": AsyncImportJob.TEMPLATE_PLANNING_MASTER_SCHEDULE,
                "farm_id": self.strict_farm.id,
                "file": upload,
            },
            format="multipart",
        )
        self.assertEqual(upload_response.status_code, status.HTTP_201_CREATED)
        job_id = upload_response.data["id"]

        validate_response = self.client.post(
            f"/api/v1/import-jobs/{job_id}/validate/",
            {},
            format="json",
        )
        self.assertEqual(validate_response.status_code, status.HTTP_200_OK)
        self.assertEqual(validate_response.data["status"], AsyncImportJob.STATUS_APPROVED_FOR_APPLY)

        apply_response = self.client.post(f"/api/v1/import-jobs/{job_id}/apply/", {}, format="json")
        self.assertEqual(apply_response.status_code, status.HTTP_200_OK)
        self.assertEqual(apply_response.data["status"], AsyncImportJob.STATUS_APPLIED)

        created_plan = CropPlan.objects.get(farm=self.strict_farm, name="خطة تشغيل ربيعية")
        self.assertEqual(created_plan.crop_id, self.crop.id)
        self.assertEqual(created_plan.season_id, self.season.id)
        self.assertTrue(created_plan.plan_locations.filter(location=self.strict_location).exists())

    def test_planning_structure_import_is_idempotent(self):
        upload = self._build_upload(
            farm=self.strict_farm,
            template_code=AsyncImportJob.TEMPLATE_PLANNING_CROP_PLAN_STRUCTURE,
            crop_plan_id=self.strict_plan.id,
            row_values=[
                str(self.task.id),
                self.task.name,
                "2026-03-10",
                "5.5",
                "General",
                "صف تشغيلي",
            ],
        )

        upload_response = self.client.post(
            "/api/v1/import-jobs/upload/",
            {
                "template_code": AsyncImportJob.TEMPLATE_PLANNING_CROP_PLAN_STRUCTURE,
                "farm_id": self.strict_farm.id,
                "crop_plan_id": self.strict_plan.id,
                "file": upload,
            },
            format="multipart",
        )
        self.assertEqual(upload_response.status_code, status.HTTP_201_CREATED)
        job_id = upload_response.data["id"]
        self.client.post(f"/api/v1/import-jobs/{job_id}/validate/", {}, format="json")
        apply_response = self.client.post(f"/api/v1/import-jobs/{job_id}/apply/", {}, format="json")
        self.assertEqual(apply_response.status_code, status.HTTP_200_OK)

        activity = PlannedActivity.objects.get(
            crop_plan=self.strict_plan,
            task=self.task,
            planned_date="2026-03-10",
        )
        self.assertEqual(activity.estimated_hours, Decimal("5.5"))

        duplicate_upload = self._build_upload(
            farm=self.strict_farm,
            template_code=AsyncImportJob.TEMPLATE_PLANNING_CROP_PLAN_STRUCTURE,
            crop_plan_id=self.strict_plan.id,
            row_values=[
                str(self.task.id),
                self.task.name,
                "2026-03-10",
                "6.0",
                "General",
                "تحديث لاحق",
            ],
        )
        duplicate_response = self.client.post(
            "/api/v1/import-jobs/upload/",
            {
                "template_code": AsyncImportJob.TEMPLATE_PLANNING_CROP_PLAN_STRUCTURE,
                "farm_id": self.strict_farm.id,
                "crop_plan_id": self.strict_plan.id,
                "file": duplicate_upload,
            },
            format="multipart",
        )
        second_job_id = duplicate_response.data["id"]
        self.client.post(f"/api/v1/import-jobs/{second_job_id}/validate/", {}, format="json")
        self.client.post(f"/api/v1/import-jobs/{second_job_id}/apply/", {}, format="json")

        activity.refresh_from_db()
        self.assertEqual(activity.estimated_hours, Decimal("6.0"))
        self.assertEqual(
            PlannedActivity.objects.filter(
                crop_plan=self.strict_plan,
                task=self.task,
                planned_date="2026-03-10",
            ).count(),
            1,
        )

    def test_planning_budget_import_updates_budget_lines(self):
        upload = self._build_upload(
            farm=self.strict_farm,
            template_code=AsyncImportJob.TEMPLATE_PLANNING_CROP_PLAN_BUDGET,
            crop_plan_id=self.strict_plan.id,
            row_values=[
                str(self.task.id),
                self.task.name,
                CropPlanBudgetLine.CATEGORY_MATERIALS,
                "2",
                "kg",
                "50",
                "100",
                "YER",
            ],
        )

        upload_response = self.client.post(
            "/api/v1/import-jobs/upload/",
            {
                "template_code": AsyncImportJob.TEMPLATE_PLANNING_CROP_PLAN_BUDGET,
                "farm_id": self.strict_farm.id,
                "crop_plan_id": self.strict_plan.id,
                "file": upload,
            },
            format="multipart",
        )
        self.assertEqual(upload_response.status_code, status.HTTP_201_CREATED)

        job_id = upload_response.data["id"]
        self.client.post(f"/api/v1/import-jobs/{job_id}/validate/", {}, format="json")
        apply_response = self.client.post(f"/api/v1/import-jobs/{job_id}/apply/", {}, format="json")
        self.assertEqual(apply_response.status_code, status.HTTP_200_OK)

        budget_line = CropPlanBudgetLine.objects.get(
            crop_plan=self.strict_plan,
            task=self.task,
            category=CropPlanBudgetLine.CATEGORY_MATERIALS,
        )
        self.assertEqual(budget_line.total_budget, Decimal("100"))
        self.strict_plan.refresh_from_db()
        self.assertEqual(self.strict_plan.budget_total, Decimal("100"))

    def test_planning_structure_duplicate_rows_fail_validation(self):
        upload = self._build_upload(
            farm=self.strict_farm,
            template_code=AsyncImportJob.TEMPLATE_PLANNING_CROP_PLAN_STRUCTURE,
            crop_plan_id=self.strict_plan.id,
            row_values=[
                str(self.task.id),
                self.task.name,
                "2026-03-10",
                "5.5",
                "General",
                "صف مكرر",
            ],
        )
        workbook = load_workbook(BytesIO(upload.read()))
        worksheet = workbook[
            ImportExportPlatformService.TEMPLATE_DEFINITIONS[
                AsyncImportJob.TEMPLATE_PLANNING_CROP_PLAN_STRUCTURE
            ].worksheet_title
        ]
        for column_index in range(1, 7):
            worksheet.cell(
                row=3,
                column=column_index,
                value=worksheet.cell(row=2, column=column_index).value,
            )
        handle = BytesIO()
        workbook.save(handle)
        duplicate_upload = SimpleUploadedFile(
            name="duplicate-structure.xlsx",
            content=handle.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        upload_response = self.client.post(
            "/api/v1/import-jobs/upload/",
            {
                "template_code": AsyncImportJob.TEMPLATE_PLANNING_CROP_PLAN_STRUCTURE,
                "farm_id": self.strict_farm.id,
                "crop_plan_id": self.strict_plan.id,
                "file": duplicate_upload,
            },
            format="multipart",
        )
        job_id = upload_response.data["id"]
        validate_response = self.client.post(f"/api/v1/import-jobs/{job_id}/validate/", {}, format="json")

        self.assertEqual(validate_response.status_code, status.HTTP_200_OK)
        self.assertEqual(validate_response.data["validation_summary"]["errors"], 1)
        self.assertEqual(validate_response.data["preview_rows"][1]["severity"], "error")
